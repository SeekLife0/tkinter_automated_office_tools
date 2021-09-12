# coding: utf-8
import sys
import os
import win32com.client as win32
import time
import xlrd
import xlwt
import pythoncom
import openpyxl
from xlutils.copy import copy
reload(sys)
sys.setdefaultencoding('utf-8')

filepaths = []                                  # 初始化列表用来
#获得文件夹下所有文件路径
def all_files_path(rootDir):
    for root, dirs, files in os.walk(rootDir):     # 分别代表根目录、文件夹、文件
        for file in files:                         # 遍历文件
            file_path = os.path.join(root, file)   # 获取文件绝对路径
            filepaths.append(file_path)            # 将文件路径添加进列表
        for dir in dirs:                           # 遍历目录下的子目录
            dir_path = os.path.join(root, dir)     # 获取子目录路径
            all_files_path(dir_path)               # 递归调用

#版本2自己返回集合
def all_files_path_WX(rootDir):
    filepaths = [] #临时集合
    for root, dirs, files in os.walk(rootDir):     # 分别代表根目录、文件夹、文件
        for file in files:                         # 遍历文件
            file_path = os.path.join(root, file)   # 获取文件绝对路径
            filepaths.append(file_path)            # 将文件路径添加进列表
        for dir in dirs:                           # 遍历目录下的子目录
            dir_path = os.path.join(root, dir)     # 获取子目录路径
            all_files_path(dir_path)               # 递归调用
    return filepaths                               # 返回文件的绝对路径

#读取路径下的所有文件名
def read_file_name(file_dir):
    L = []
    for root,dirs,files in os.walk(file_dir):
        for file in files:
            L.append(file)
    return L

#进行xls到xlsx的转换
def xlstoxlsx(file,filePath):
    if file.find('xls'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        file = file.decode('utf-8').encode('gbk')
        filePath1 = filePath.decode('utf-8').encode('gbk')
        wb = None
        try:
            wb = excel.Workbooks.Open(file)
            wb.SaveAs(filePath1 + 'x', FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            print filePath + 'x'
        except IOError:
            print '文件读写错误'
        finally:
            wb.Close()
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath + 'x'

#进行xlsx到xls的转换
def xlsxtoxls(file,filePath):
    if file.find('xlsx'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        file = file.decode('utf-8').encode('gbk')
        #去除末尾的x[-1:]
        # filepaths = filePath.split('.')
        # filePath1 = filepaths[0]
        filePath1 = "{}".format(filePath[:-4])
        filePath1 = filePath1.decode('utf-8').encode('gbk')
        #打开文件需要进行异常捕获和关闭文件
        wb = None
        try:
            wb = excel.Workbooks.Open(file)
            wb.SaveAs(filePath1 + '.xls', FileFormat=56)  # FileFormat = 51 is for .xlsx extension
            print filePath1 + '.xls'
        except IOError:
            print '该文件读写发生错误'
        finally:
            wb.Close()
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath1 + '.xls'

#写入到execel表格
#先打开原表然后复制一份最后另存为
def wExecel(path,row,col,startCol,values):
    #这里需要对路径文件格式进行判断一个if语句
    if path.find('xlsx') != -1:                    #把xlsx转为xls，会把原路径的xlsx文件替换
        xlsxPath = path.decode('utf-8').encode('gbk')
        print xlsxPath
    else:
        print '文件格式为：xls'
        xlsxPath = xlstoxlsx(path,path)
        #转换完成之后需要把原来xlsx这个文件删除
        path = path.decode('utf-8').encode('gbk')
        os.remove(path)
        print '文件已删除'
        print xlsxPath
    wb = None
    try:
        wb = openpyxl.load_workbook(xlsxPath)
        sheetnames = wb.get_sheet_names()
        # sheet = wb.get_sheet_by_name(sheetnames[0])
        j = 0
        for sheetname in sheetnames:
            sheet = wb.get_sheet_by_name(sheetname)  # 取sheet表
            # i = col
            for num in range(0,len(values)-(col-1),col):
                for a in range(0,col):
                    value = str(values[num+a])
                    # value = value.decode('utf-8').encode('gbk')
                    sheet.cell(row,startCol+a).value = value  # 姓名
                row += 1
            j += 1
        #这里保存的路径需要修改，只需要路径名
        #path路径以点分隔把文件格式后缀删除加上重命名即可
        print path
        wb.save(path)
    except IOError:
        print 'IO读写错误'

#因为涉及到一个文件可能存在多个表的问题所以要以表为单位进行内容抓取
def rwExecel(Rpath,eList_col,eList_entry):
    # global list_col
    xlsPath = Rpath
    xls = None
    #这里需要对路径文件格式进行判断一个if语句
    if 'xlsx' in Rpath:                    #把xlsx转为xls，会把原路径的xlsx文件替换
        print '文件格式为：xlsx'
        xlsPath = xlsxtoxls(Rpath,Rpath)
        #转换完成之后需要把原来xlsx这个文件删除
        path = Rpath.decode('utf-8').encode('gbk')
        os.remove(path)
        print '文件已删除'
        print xlsPath
        # xls = xlrd.open_workbook(xlsPath)
    else:
        xlsPath = xlsPath.decode('utf-8').encode('gbk')
        print xlsPath
    xls = xlrd.open_workbook(xlsPath)
    tables = xls.sheets()             #获取文件所有表格这样就不需要表的标号和名称了
    sheet_name = xls.sheet_names()
    print sheet_name
    for table in tables:              #如何判断一个表是否有内容
        if table.nrows == 0:          #查看表的有效函数是否为0
            continue                  #直接进入下一个循环
        #通过遍历elist_entry来获得对应坐标,每三个为一个单位获取
        for num in range(0,len(eList_entry)-2,3):
            #只读入有效坐标只有行或列或者全没有不读入直接跳入下一个循环
            if (eList_entry[num+1].get()!='' and eList_entry[num+2].get()!='') and (eList_entry[num+1].get()!=None and eList_entry[num+2].get()!=None):
                r = int(eList_entry[num+1].get())   #第二个是行
                c = int(eList_entry[num+2].get())   #第三个是列
                eList_col.append(table.cell_value(r,c))
                print '每次爬取内容' + str(table.cell_value(r,c))
            else:
                continue
        # wExecel(Wpath,eRow,eCol,eList_col) #写入的文件路径
        # del eList_col[:]  #每次执行完一行清空列表
        # eRow = eRow + 1
    return eList_col  #返回装载所有多表信息的列表