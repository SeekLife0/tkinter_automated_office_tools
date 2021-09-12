# coding: utf-8
import sys
import os
import win32com.client as win32
import time
import xlrd
import xlwt
import openpyxl
from xlutils.copy import copy
reload(sys)
sys.setdefaultencoding('utf-8')
import pythoncom


filepaths = []                                  # 初始化列表用来
#获得文件夹下所有文件路径
def all_files_path(rootDir):
    for root, dirs, files in os.walk(rootDir):     # 分别代表根目录、文件夹、文件
        for file in files:                         # 遍历文件
            file_path = os.path.join(root, file)   # 获取文件绝对路径
            filepaths.append(file_path)            # 将文件路径添加进列表
        for dir in dirs:                           # 遍历目录下的子目录
            dir_path = os.path.join(root, dir)     # 获取子目录路径
            all_files_path(dir_path)               # 递归调用

#读取路径下的所有文件名
def read_file_name(file_dir):
    L = []
    for root,dirs,files in os.walk(file_dir):
        for file in files:
            L.append(file)
    return L

#进行xls到xlsx的转换
def xlstoxlsx(file,filePath):
    if file.find('xls'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        file = file.decode('utf-8').encode('gbk')
        filePath1 = filePath.decode('utf-8').encode('gbk')
        wb = None
        try:
            wb = excel.Workbooks.Open(file)
            wb.SaveAs(filePath1 + 'x', FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            print filePath + 'x'
        except IOError as e:
            print e
        finally:
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath + 'x'

#进行xlsx到xls的转换
def xlsxtoxls(file,filePath):
    if file.find('xlsx'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        file = file.decode('utf-8').encode('gbk')
        #去除末尾的x[-1:]
        # filepaths = filePath.split('.')
        # filePath1 = filepaths[0]
        filePath1 = "{}".format(filePath[:-4])
        filePath1 = filePath1.decode('utf-8').encode('gbk')
        #打开文件需要进行异常捕获和关闭文件
        wb = None
        try:
            wb = excel.Workbooks.Open(file)
            wb.SaveAs(filePath1 + 'xls', FileFormat=56)  # FileFormat = 51 is for .xlsx extension
            print filePath1 + 'xls'
        except IOError as e:
            print e
        finally:
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath1 + 'xls'

#写入到execel表格
#先打开原表然后复制一份最后另存为
def wExecel1(path,values,reName):
    oldWb = xlrd.open_workbook(path,formatting_info=True);  # 先打开已存在的表
    newWb = copy(oldWb)  # 复制整个文件,转未xlwt对象这样就可以进行追加写入操作
    #获取所有表
    tables = oldWb.sheets()
    i = 0
    for table in tables:
        if table.nrows == 0:     #查看表的有效函数是否为0
            continue            #直接进入下一个循环
        newWbs = newWb.get_sheet(i);  # 取sheet表
        i += 1
        for num in range(0,len(values) - 2,3):
            value = values[num]
            row = values[num+1]
            col = values[num+2]
            print '要写入的坐标行:'+str(row)+'列：'+str(col)
            #写入的时候添加对应格式,比如居中显示
            style = xlwt.easyxf('font:height 240, color-index black, bold off;align: wrap on, vert centre, horiz center');
            newWbs.write(row, col, value,style)  # 姓名
    #这里保存的路径需要修改，只需要路径名
    #path路径以点分隔把文件格式后缀删除加上重命名即可
    print path
    print '每个文件重命名的名称'+reName
    eFilePath = path.split('.')[0] + reName + '.xls'
    print '保存路径为：' + eFilePath
    newWb.save(eFilePath)

#测试写入格式操作
def wExecel2(path,values,reName):
    #设置要修改的样式
    #设置边框
    borders = xlwt.Borders()
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom = 1
    #设置字体
    font = xlwt.Font()
    # font.name = '宋体'    #不能写宋体这个中文字符，直接说ascii码无法解析
    font.height = 240
    #设置居中对齐
    alignment = xlwt.Alignment()
    alignment.horz = 0x02   #水平方向上居中对齐
    alignment.vert = 0x01   #垂直方向上居中对齐
    alignment.wrap = 1   #自动换行   1自动换行  0不自动换行
    #把要修改的内容放入到style的对象
    style = xlwt.XFStyle()
    style.borders = borders
    style.font = font
    style.alignment = alignment
    oldWb = xlrd.open_workbook(path,formatting_info=True,on_demand=True);  # 先打开已存在的表
    newWb = copy(oldWb)  # 复制整个文件,转未xlwt对象这样就可以进行追加写入操作
    #获取所有表
    tables = oldWb.sheets()
    i = 0
    for table in tables:
        if table.nrows == 0:     #查看表的有效函数是否为0
            continue            #直接进入下一个循环
        newWbs = newWb.get_sheet(i);  # 取sheet表
        i += 1
        for num in range(0,len(values) - 2,3):
            value = values[num]
            row = values[num+1]
            col = values[num+2]
            print '要写入的坐标行:'+str(row)+'列：'+str(col)
            #写入的时候添加对应格式,比如居中显示
            # style = xlwt.easyxf('font:height 240, color-index black, bold off;align: wrap on, vert centre, horiz center');
            newWbs.write(row, col, value,style)  # 姓名
    #这里保存的路径需要修改，只需要路径名
    #path路径以点分隔把文件格式后缀删除加上重命名即可
    print path
    print '每个文件重命名的名称'+reName
    eFilePath = path.split('.')[0] + reName + '.xls'
    print '保存路径为：' + eFilePath
    #add_sheet方法只是单纯的创建了一个空白表格
    # newWb.add_sheet('sheet11',cell_overwrite_ok = True)
    newWb.save(eFilePath)

#使用openpyxl写入到execel表格
def wExecel3(path,values,reName):
    #这里需要对路径文件格式进行判断一个if语句
    if 'xlsx' in path:                    #把xlsx转为xls，会把原路径的xlsx文件替换
        xlsxPath = path.decode('utf-8').encode('gbk')
        print xlsxPath
    else:
        print '文件格式为：xlsx'
        xlsxPath = xlstoxlsx(path,path)
        #转换完成之后需要把原来xlsx这个文件删除
        path = path.decode('utf-8').encode('gbk')
        os.remove(path)
        print '文件已删除'
        print xlsxPath
    wb = None
    # try:
    wb = openpyxl.load_workbook(xlsxPath)
    sheetnames = wb.get_sheet_names()
    # sheet = wb.get_sheet_by_name(sheetnames[0])
    i = 0
    for sheetname in sheetnames:
        sheet = wb.get_sheet_by_name(sheetname)  # 取sheet表
        i += 1
        for num in range(0,len(values) - 2,3):
                value = values[num]
                row = values[num+1]
                col = values[num+2]
                print '写入的值：'+ str(value) + 'row:' + str(row) + 'col:' + str(col)
                sheet.cell(row,col).value = value
    #这里保存的路径需要修改，只需要路径名
    #path路径以点分隔把文件格式后缀删除加上重命名即可
    print path
    print '每个文件重命名的名称'+reName
    eFilePath =  "{}".format(path[:-7]) + reName  + '.xlsx'   #命名文件必须为两个字符
    print '保存路径为：' + eFilePath
    wb.save(eFilePath)
    # except IOError as e:
    #     print e

    # oldWb = xlrd.open_workbook(path,formatting_info=True);  # 先打开已存在的表
    # newWb = copy(oldWb)  # 复制整个文件,转未xlwt对象这样就可以进行追加写入操作
    # #获取所有表
    # tables = oldWb.sheets()
    # i = 0
    # for table in tables:
    #     if table.nrows == 0:     #查看表的有效函数是否为0
    #         continue            #直接进入下一个循环
    #     newWbs = newWb.get_sheet(i);  # 取sheet表
    #     i += 1
    #     for num in range(0,len(values) - 2,3):
    #         value = values[num]
    #         row = values[num+1]
    #         col = values[num+2]
    #         print '要写入的坐标行:'+str(row)+'列：'+str(col)
    #         #写入的时候添加对应格式,比如居中显示
    #         style = xlwt.easyxf('font:height 240, color-index black, bold off;align: wrap on, vert centre, horiz center');
    #         newWbs.write(row, col, value,style)  # 姓名
    # #这里保存的路径需要修改，只需要路径名
    # #path路径以点分隔把文件格式后缀删除加上重命名即可
    # print path
    # print '每个文件重命名的名称'+reName
    # eFilePath = path.split('.')[0] + reName + '.xls'
    # print '保存路径为：' + eFilePath
    # newWb.save(eFilePath)

#因为涉及到一个文件可能存在多个表的问题所以要以表为单位进行内容抓去
def rwExecel_B(Rpath,Wpath,rowValues,reNameCol,eList_entry):
    list_data = []
    xlsPath = Rpath
    xls = None
    reNameC = ''
    #这里需要对路径文件格式进行判断一个if语句
    if Rpath.find('.xlsx') != -1:                    #把xlsx转为xls，会把原路径的xlsx文件替换
        print '文件格式为：xlsx'
        xlsPath = xlsxtoxls(Rpath,Rpath)
        #转换完成之后需要把原来xlsx这个文件删除
        path = Rpath.decode('utf-8').encode('gbk')
        os.remove(path)
        print '文件已删除'
        print xlsPath
    else:
        print '要读取的文件是xls无需转化'
        xlsPath = xlsPath.decode('utf-8').encode('gbk')
        print xlsPath
    # try:
    xls = xlrd.open_workbook(xlsPath)
    tables = xls.sheets()         #获取文件所有表格这样就不需要表的标号和名称了
    sheet_name = xls.sheet_names()
    print sheet_name
    for table in tables:         #如何判断一个表是否有内容
        if table.nrows == 0:     #查看表的有效函数是否为0
            continue            #直接进入下一个循环
        #通过遍历elist_entry来获得对应坐标,每三个为一个单位获取
        for r in rowValues:
            #遍历每一列
            for num in range(0,len(eList_entry)-3,4):
                if (eList_entry[num+1].get()!='' and eList_entry[num+2].get()!='' and eList_entry[num+3].get()!='') and (eList_entry[num+1].get()!=None and eList_entry[num+2].get()!=None and eList_entry[num+3].get()!=None):
                    #每四个为一组，列名，列数，复制到坐标
                    #1获取所有的列,第二个是列
                    c = int(eList_entry[num+1].get())
                    # print reNameCol
                    if c == int(reNameCol):
                        print '重命名的行列' + str(r) + str(c)
                        list_data.append(table.cell_value(r,c))
                        reNameC = table.cell_value(r,c)
                        print '重命名：' + reNameC
                    else:
                        #获取坐标对应的内容，这里是一整行的内容
                        list_data.append(table.cell_value(r,c))
                    print '每次爬取内容' + str(table.cell_value(r,c))   #打印内容进行测试
                    #获取对应要填写的坐标
                    rAim = int(eList_entry[num+2].get())  #第三个是坐标行
                    cAim = int(eList_entry[num+3].get())   #第四个是坐标列
                    list_data.append(rAim)
                    list_data.append(cAim)
                    #把所有列遍历完毕，再来进行写入减少io操作
                else:
                    continue
            wExecel3(Wpath,list_data,reNameC) #写入的文件路径
            del list_data[:]  #每次执行完一行清空列表，情况的是列的数据
    # except IOError as e:
    #     print e